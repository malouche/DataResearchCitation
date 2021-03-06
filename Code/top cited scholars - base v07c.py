# Databricks notebook source
# MAGIC %md # init functions

# COMMAND ----------

import sys
try:
  df_ani
  minyear # only applies to the column npY1Y3
  mincityear
  minyear_overall # will limit entire citation/publication metric data to only include papers since x.
  maxyear
  includepp # include or exclude pp in the publication counts (pp citations are excluded by default)
  ani
except NameError:
  print("this notebook should not be run directly. It needs parameters set and then called via a %run command")
  sys.exit


# COMMAND ----------

def file_exists(path):
  if path[:5] == "/dbfs":
    import os
    return os.path.exists(path)
  else:
    try:
      dbutils.fs.ls(path)
      return True
    except Exception as e:
      if 'java.io.FileNotFoundException' in str(e):
        return False
      else:
        raise

# COMMAND ----------

from pyspark.sql.types import *
from pyspark.sql import functions as func
from pyspark.sql import Window
from pyspark.sql import Row

# for printing labels on the result columns:
Y1=minyear[2:]
Y2=mincityear[2:]
Y3=maxyear[2:]
Y4='None' if minyear_overall is None else minyear_overall
# keep empty file name Y4_name_postfix if None.
Y4_name_postfix='' if minyear_overall is None else f'_pubs_since_{Y4}'

if minyear_overall is not None:
  if minyear_overall > mincityear:
    print("minyear_overall should be less than mincityear")
    sys.exit

csvOutFileName_s1=basePath_project+'Table-S1_'+ani['stamp']+'_'+minyear+'_'+mincityear+'_'+maxyear+'_'+Y4+'_pp'+str(includepp)
csvOutFileName_s3=basePath_project+'Table-S3_'+ani['stamp']+'_'+minyear+'_'+mincityear+'_'+maxyear+'_'+Y4+'_pp'+str(includepp)

if discontinued_sources is None:
  discontinued_sources=[]


# COMMAND ----------

# MAGIC %md # author table

# COMMAND ----------

# for each author in Scopus, get their latest record and select the name from that record.
# we will use this in the end to label each author record for which we create stats.


tbn_df_ani_auth_name="temp_df_ani_auth_name"
savepath_auth_name=basePath_project+"cache/"+ani['stamp']+"/"+tbn_df_ani_auth_name
if file_exists(savepath_auth_name):
  print("path already exists: "+savepath_auth_name)
else:
  print ("Generating "+savepath_auth_name)
  df_ani_auth_name_t=(
    df_ani
    .withColumn('Au',func.explode('Au'))
    .withColumn('auid',func.col('Au.auid').cast('long'))
    .orderBy('datesort',Ascending=False)
    #.filter('Au.given_name_pn IS NOT NULL')
    # no need to filter if we order and prefer the presence of the given_name:
    .withColumn(
      'recRank',
      func.rank().over(
        Window
        .partitionBy('auid')
        .orderBy(
          func.expr('IF(Au.given_name_pn IS NULL,0,1)').desc(),
          func.desc('datesort'),
          func.desc('Eid'),
          func.desc('Au.authorseq') # have issues with records with multiple auid repeats on same record. this breaks those ties.
        )
      )
    )
    .withColumn('lastPubYear',func.max(sort_pub_year).over(Window.partitionBy('auid')))
    .withColumn('firstPubYear',func.min(sort_pub_year).over(Window.partitionBy('auid')))
    .filter('recRank=1')
    .select('auid',func.col('Eid').alias('LastEidName'),func.coalesce(func.expr('CONCAT(Au.surname_pn,", ",IFNULL(Au.given_name_pn,Au.initials_pn))'),func.col('Au.indexed_name_pn'),func.col('Au.indexed_name')).alias('authfull'),'lastPubYear','firstPubYear')
  )
  df_ani_auth_name_t.repartition(200).write.mode("overwrite").format("parquet").save(savepath_auth_name)

# COMMAND ----------

# MAGIC %md # citation counts

# COMMAND ----------

# citation counts (mincityear+)
def get_df_mincityear_onw_cit(df_ani):
  return (
    df_ani
    .filter(sort_pub_year+' >= '+mincityear)
    # below filter only works from ICSR-Lab-data-v005+ is needed to exclude pre-print citations.
    .filter(func.arrays_overlap('dbcollections',func.array(func.lit('SCOPUS'),func.lit('MEDL'))))
    .withColumn('references_u',func.array_distinct('references'))
    .select(
      func.col('Eid').alias('CitingEid'),
      func.explode('references_u').alias('Eid'),
      func.when(func.col('source.srcid').isin(discontinued_sources),func.lit(int(1))).otherwise(func.lit(int(0))).alias('isDiscontinuedCiting'),
      func.col('Au.auid').cast('array<long>').alias('CitingAuids')
    )
    .join(
      df_ani.select(
        'Eid',
        func.col('Au.auid').cast('array<long>').alias('CitedAuids')
      ),["Eid"]
    )
    .withColumn('overLappingAuthors',func.size(func.array_intersect('CitingAuids','CitedAuids')))
    .select(
      "CitingEid",
      "Eid",
      'isDiscontinuedCiting',
      func.expr("IF(overLappingAuthors>0,1,0)").alias('isSelfCitation'),
      func.expr("IF(overLappingAuthors>0,NULL,CitingEid)").alias('CitingEidNonSelf'),
    )
    .groupBy('Eid')
    .agg(
      func.count('*').alias('CitationCount'),
      func.sum('isSelfCitation').alias('SelfCitationCount'),
      (func.count('*')-func.sum('isSelfCitation')).alias('CitationCountNonSelf'),
      func.collect_list('CitingEid').alias('CitingEids'),
      func.collect_list('CitingEidNonSelf').alias('CitingEidsNonSelf'),
      func.sum("isDiscontinuedCiting").alias('CitationCountFromDiscontinuedSources')
    )
  )

# COMMAND ----------

# MAGIC %md # windows

# COMMAND ----------

# window definitions used in next command to calculate the volume and rank of the (sub)fields per author to assign the best subject per author.
w_au=Window.partitionBy('auid')
w_au_sf=Window.partitionBy('auid','subfield')
w_sf=Window.partitionBy('subfield')
# this window won't have ties. That way the dense_rank can be used to de-duplicate the results 
# (i.e. all those ranked #1 will be the same item, no risk of two items being ranked #1)
# note also the first subfield IS NULL: designed to put the count of null subfields to the END of the rank.
w_au_sort_sf_auth_overall=Window.partitionBy('auid').orderBy(
  func.expr('subfield IS NULL').asc(),
  func.desc('subfield_count_thisauth'),
  func.asc('subfield_count_overall'),
  func.asc('subfield')
)
w_au_f=Window.partitionBy('auid','field')
w_f=Window.partitionBy('field')
# see comment about ties above.
w_au_sort_f_auth_overall=Window.partitionBy('auid').orderBy(
  func.expr('field IS NULL').asc(),
  func.desc('field_count_thisauth'),
  func.asc('field_count_overall'),
  func.asc('field')
)

# COMMAND ----------

# MAGIC %md # main author-citations table

# COMMAND ----------

# calculate the metric per author in Scopus.
# Note: ns_ and ws_ prefixes as found in the code, indicate NonSelf and WithSelf citation counts used.

# list of metrics:

# npY1Y3 # papers minyear-maxyear
# firstyr year of first publication
# lastyr year of most recent publication
# ncY2Y3 total cites mincityear-maxyear
# ncY2Y3_cp number of citing papers (unique) citing ncY2Y3
# hY3 h-index as of end-maxyear
# hmY3 hm-index as of end-maxyear
# nps number of single authored papers
# ncs total cites to single authored papers
# npsf number of single+first authored papers
# ncsf total cites to single+first authored papers
# npsfl number of single+first+last authored papers
# ncsfl total cites to single+first+last authored papers

wns = (Window.partitionBy('auid').orderBy(func.desc('CitationCountNonSelf'),'Eid'))
wws = (Window.partitionBy('auid').orderBy(func.desc('CitationCount'),'Eid'))
wsm = (Window.partitionBy('auid','subfield'))
wsm_a = (Window.partitionBy('subfield'))

# store this as a table if it doesn't already exist.
tbn_df_agg_count="temp_df_agg_count"
savepath_agg_count=basePath_project+"cache/"+ani['stamp']+"/"+tbn_df_agg_count+'_'+minyear+'_'+mincityear+'_'+maxyear+'_'+Y4+'_pp'+str(includepp)
if file_exists(savepath_agg_count):
  print("path already exists: "+savepath_agg_count)
else:
  print ("Generating "+savepath_agg_count)
  df_agg_count_t=(
    df_ani
    # only include pp in counts if selected
    .filter(func.lit(True) if includepp else func.arrays_overlap('dbcollections',func.array(func.lit('SCOPUS'),func.lit('MEDL'))))
    
    .drop('CitationCount') # don't use pre-provided citation count.
    #.filter('publication_type IN ("ar","cp","re")')
    .filter(sort_pub_year+' <= '+maxyear)
    .join(get_df_mincityear_onw_cit(df_ani),["Eid"],"LEFT_OUTER")
    .na.fill({'CitationCountNonSelf':0,'CitationCount':0,'CitationCountFromDiscontinuedSources':0})
    .select(
      'Eid',
      sort_pub_year,
      func.when(func.col('source.srcid').isin(discontinued_sources),func.lit(int(1))).otherwise(func.lit(int(0))).alias('isDiscontinuedSource'),
      func.size('Au').alias('n_authors'),
      func.explode('Au').alias('Au'),
      'CitationCountNonSelf','CitationCount','CitingEidsNonSelf','CitingEids',func.col('source.srcid').alias('srcid'),'publication_type','CitationCountFromDiscontinuedSources'
    )
    .withColumn('auid',func.col('Au.auid').cast('long'))
    .withColumn('Authorseq',func.col('Au.Authorseq')).drop('Au')
    .filter('auid > 4 AND auid IS NOT NULL')
    
    # join ScienceMetrix to count volume per subfield code
    .join(df_smc,["Eid"],"LEFT_OUTER")
    .withColumn('subfield',func.expr('IF(publication_type IN ("ar","cp","re"),subfield,NULL)')) # assign field only if ar/cp/re
    .withColumn('totalDocCountAuthor_alltime',func.count('*').over(w_au))
    .withColumn('arcpreDocCountAuthor_alltime',func.count(func.expr('IF(publication_type IN ("ar","cp","re"),Eid,NULL)')).over(w_au))
    .filter('arcpreDocCountAuthor_alltime>=5') # only authors with >= 5 ar/cp/re's 
    .withColumn('SMFCMdocCountAuthor',func.count('field').over(w_au)) # number of documents the author has mapped to a S-M; 
    # calculate ranks for 176-level subfield per author
    .withColumn('subfield_count_thisauth',func.count('*').over(w_au_sf))
    .withColumn('subfield_count_overall',func.count('*').over(w_sf))
    .withColumn('subfieldRank',func.dense_rank().over(w_au_sort_sf_auth_overall))
    .withColumn(
      "subfield_tuple",
      func.struct(
        'subfieldRank',
        'subfield',
        'subfield_count_thisauth',
        'subfield_count_overall',
        'SMFCMdocCountAuthor',
        func.expr('subfield_count_thisauth/SMFCMdocCountAuthor').alias('subfieldFrac')
      )
    )
    # same for 22-level fields
    .withColumn('field_count_thisauth',func.count('*').over(w_au_f))
    .withColumn('field_count_overall',func.count('*').over(w_f))
    .withColumn('fieldRank',func.dense_rank().over(w_au_sort_f_auth_overall))
    .withColumn(
      "field_tuple",
      func.struct(
        'fieldRank',
        'field',
        'field_count_thisauth',
        'field_count_overall',
        'SMFCMdocCountAuthor',
        func.expr('field_count_thisauth/SMFCMdocCountAuthor').alias('fieldFrac')
      )
    )

    # ranks and sums needed to calculate h and hm index
    .withColumn('ns_r_eff',func.sum(1/func.col('n_authors')).over(wns.rangeBetween(Window.unboundedPreceding, 0)))
    .withColumn('ns_r',func.rank().over(wns))
    .withColumn('ws_r_eff',func.sum(1/func.col('n_authors')).over(wws.rangeBetween(Window.unboundedPreceding, 0)))
    .withColumn('ws_r',func.rank().over(wws))

    .filter(func.lit(True) if minyear_overall is None else (sort_pub_year+' >= '+minyear_overall))#v07b filter: a lower year filter. only apply if not None.
    
    .groupBy('auid')
    .agg(
      func.sort_array(func.collect_set("subfield_tuple"),True).alias("subFields"),
      func.sort_array(func.collect_set("field_tuple"),True).alias("Fields"),
      func.count('*').alias('np'),
      func.sum(func.expr('IF('+sort_pub_year+' BETWEEN '+minyear+' AND '+maxyear+',1,0)')).alias('npY1Y3'),
      # no longer capture first/last here; we want to get those values from the full database and therefore collect them with the author names dataframe (where we also get the last known full prefereed name)
      #func.min(sort_pub_year).alias('firstyr'),
      #func.max(sort_pub_year).alias('lastyr'),
      
      # number of cited papers.
      func.sum(func.expr('IF('+sort_pub_year+' BETWEEN '+minyear+' AND '+maxyear+',IF(CitationCountNonSelf>0,1,0),0)')).alias('ns_npcY1Y3'),
      func.sum(func.expr('IF('+sort_pub_year+' BETWEEN '+minyear+' AND '+maxyear+',IF(CitationCount>0,1,0),0)')).alias('ws_npcY1Y3'),
      # number of cited papers.
      func.sum(func.expr('IF(CitationCountNonSelf>0,1,0)')).alias('ns_npc'),
      func.sum(func.expr('IF(CitationCount>0,1,0)')).alias('ws_npc'),
      
      func.sum('CitationCountNonSelf').alias('ns_ncY2Y3'),
      func.size(func.array_distinct(func.flatten(func.collect_list('CitingEidsNonSelf')))).alias('ns_ncY2Y3_cp'),
      func.max(func.expr('IF(ns_r<=CitationCountNonSelf,ns_r,0)')).alias('ns_hY3'),
      func.max(func.expr('IF(ns_r_eff<=CitationCountNonSelf,ns_r_eff,0)')).alias('ns_hmY3'),
      func.sum(func.expr('IF(n_authors=1,1,0)')).alias('ns_nps'),
      func.sum(func.expr('IF(n_authors=1,CitationCountNonSelf,0)')).alias('ns_ncs'),
      func.sum(func.expr('IF(n_authors=1 OR Authorseq=1,1,0)')).alias('ns_npsf'),
      func.sum(func.expr('IF(n_authors=1 OR Authorseq=1,CitationCountNonSelf,0)')).alias('ns_ncsf'),
      func.sum(func.expr('IF(n_authors=1 OR Authorseq=1 OR Authorseq=n_authors,1,0)')).alias('ns_npsfl'),
      func.sum(func.expr('IF(n_authors=1 OR Authorseq=1 OR Authorseq=n_authors,CitationCountNonSelf,0)')).alias('ns_ncsfl'),

      func.sum('CitationCount').alias('ws_ncY2Y3'),
      func.size(func.array_distinct(func.flatten(func.collect_list('CitingEids')))).alias('ws_ncY2Y3_cp'),
      func.max(func.expr('IF(ws_r<=CitationCount,ws_r,0)')).alias('ws_hY3'),
      func.max(func.expr('IF(ws_r_eff<=CitationCount,ws_r_eff,0)')).alias('ws_hmY3'),
      func.sum(func.expr('IF(n_authors=1,1,0)')).alias('ws_nps'),
      func.sum(func.expr('IF(n_authors=1,CitationCount,0)')).alias('ws_ncs'),
      func.sum(func.expr('IF(n_authors=1 OR Authorseq=1,1,0)')).alias('ws_npsf'),
      func.sum(func.expr('IF(n_authors=1 OR Authorseq=1,CitationCount,0)')).alias('ws_ncsf'),
      func.sum(func.expr('IF(n_authors=1 OR Authorseq=1 OR Authorseq=n_authors,1,0)')).alias('ws_npsfl'),
      func.sum(func.expr('IF(n_authors=1 OR Authorseq=1 OR Authorseq=n_authors,CitationCount,0)')).alias('ws_ncsfl'),

      func.sum(func.expr('IF('+sort_pub_year+' BETWEEN '+minyear+' AND '+maxyear+',isDiscontinuedSource,0)')).alias('npY1Y3_d'),
      func.sum('isDiscontinuedSource').alias('np_d'),
      func.sum('CitationCountFromDiscontinuedSources').alias('ws_ncY2Y3_d')
    )
  )
  df_agg_count_t.repartition(200).write.mode("overwrite").format("parquet").save(savepath_agg_count)

def get_df_agg_count_ln(savepath_agg_count):
  df_agg_count=spark.read.format("parquet").load(savepath_agg_count)
  # more metrics: log scales of the calculated metrics.
  # lnc	ln(ncY2Y3+1)
  # lh	ln(hY3+1)
  # lhm	ln(hmY3+1)
  # lns	ln(ncs+1)
  # lnsf	ln(ncsf+1)
  # lnsfl	ln(ncsfl+1)
  return (
    df_agg_count
    .withColumn('ns_lnc',func.log(func.expr('ns_ncY2Y3+1')))
    .withColumn('ns_lh',func.log(func.expr('ns_hY3+1')))
    .withColumn('ns_lhm',func.log(func.expr('ns_hmY3+1')))
    .withColumn('ns_lns',func.log(func.expr('ns_ncs+1')))
    .withColumn('ns_lnsf',func.log(func.expr('ns_ncsf+1')))
    .withColumn('ns_lnsfl',func.log(func.expr('ns_ncsfl+1')))

    .withColumn('ws_lnc',func.log(func.expr('ws_ncY2Y3+1')))
    .withColumn('ws_lh',func.log(func.expr('ws_hY3+1')))
    .withColumn('ws_lhm',func.log(func.expr('ws_hmY3+1')))
    .withColumn('ws_lns',func.log(func.expr('ws_ncs+1')))
    .withColumn('ws_lnsf',func.log(func.expr('ws_ncsf+1')))
    .withColumn('ws_lnsfl',func.log(func.expr('ws_ncsfl+1')))
  )  


# and to complement further, the MAX of the log functions above (aggregate across the DB)
# flnc	lnc/max(lnc)
# flh	lh/max(lh)
# flhm	lhm/max(lhm)
# flns	lns/max(lns)
# flnsf	lnsf/max(lnsf)
# flnsfl	lnsfl/max(lnsfl)
# c	composite: sum of six fractions (flnc+flh+flhm+flns+flnsf+flnsfl)

tbn_df_agg_count_ln_max="temp_df_agg_count_ln_max"
savepath_agg_count_ln_max=basePath_project+"cache/"+ani['stamp']+"/"+tbn_df_agg_count_ln_max+'_'+minyear+'_'+mincityear+'_'+maxyear+'_'+Y4+'_pp'+str(includepp)
if file_exists(savepath_agg_count_ln_max):
  print("path already exists: "+savepath_agg_count_ln_max)
else:
  df_agg_count_ln_max_t=get_df_agg_count_ln(savepath_agg_count).agg(
    func.max('ns_lnc').alias('ns_maxlnc'),
    func.max('ns_lh').alias('ns_maxlh'),
    func.max('ns_lhm').alias('ns_maxlhm'),
    func.max('ns_lns').alias('ns_maxlns'),
    func.max('ns_lnsf').alias('ns_maxlnsf'),
    func.max('ns_lnsfl').alias('ns_maxlnsfl'),

    func.max('ws_lnc').alias('ws_maxlnc'),
    func.max('ws_lh').alias('ws_maxlh'),
    func.max('ws_lhm').alias('ws_maxlhm'),
    func.max('ws_lns').alias('ws_maxlns'),
    func.max('ws_lnsf').alias('ws_maxlnsf'),
    func.max('ws_lnsfl').alias('ws_maxlnsfl')  
  )
  print ("Generating "+savepath_agg_count_ln_max)
  df_agg_count_ln_max_t.repartition(200).write.mode("overwrite").format("parquet").save(savepath_agg_count_ln_max)

tbn_df_agg_count_ln_wmaxc="temp_df_agg_count_ln_wmaxc"
savepath_agg_count_ln_wmaxc=basePath_project+"cache/"+ani['stamp']+"/"+tbn_df_agg_count_ln_wmaxc+'_'+minyear+'_'+mincityear+'_'+maxyear+'_'+Y4+'_pp'+str(includepp)
if file_exists(savepath_agg_count_ln_wmaxc):
  # now bring the aggregate back into the metrics table so we can calculate the normalized score.
  print("path already exists: "+savepath_agg_count_ln_wmaxc)
else:
  print ("Generating "+savepath_agg_count_ln_wmaxc)

  df_agg_count_ln_max=spark.read.format("parquet").load(savepath_agg_count_ln_max)
  
  if not file_exists(csvOutFileName_s1+'_lnmaxvalues.csv'):
    df_agg_count_ln_max.repartition(1).write.mode("overwrite").csv(csvOutFileName_s1+'_lnmaxvalues.csv',header = 'true')

  df_agg_count_ln_wmaxc_t=(
    get_df_agg_count_ln(savepath_agg_count)
    .crossJoin(df_agg_count_ln_max)
    #.withColumn('flnc',func.expr('lnc/(max(lnc) over(Partition By true))'))
    .withColumn('ns_flnc',func.expr('ns_lnc/ns_maxlnc')).drop('ns_maxlnc')
    .withColumn('ns_flh',func.expr('ns_lh/ns_maxlh')).drop('ns_maxlh')
    .withColumn('ns_flhm',func.expr('ns_lhm/ns_maxlhm')).drop('ns_maxlhm')
    .withColumn('ns_flns',func.expr('ns_lns/ns_maxlns')).drop('ns_maxlns')
    .withColumn('ns_flnsf',func.expr('ns_lnsf/ns_maxlnsf')).drop('ns_maxlnsf')
    .withColumn('ns_flnsfl',func.expr('ns_lnsfl/ns_maxlnsfl')).drop('ns_maxlnsfl')
    # c = the composite metric
    .withColumn('ns_c',func.expr('ns_flnc+ns_flh+ns_flhm+ns_flns+ns_flnsf+ns_flnsfl'))

    .withColumn('ws_flnc',func.expr('ws_lnc/ws_maxlnc')).drop('ws_maxlnc')
    .withColumn('ws_flh',func.expr('ws_lh/ws_maxlh')).drop('ws_maxlh')
    .withColumn('ws_flhm',func.expr('ws_lhm/ws_maxlhm')).drop('ws_maxlhm')
    .withColumn('ws_flns',func.expr('ws_lns/ws_maxlns')).drop('ws_maxlns')
    .withColumn('ws_flnsf',func.expr('ws_lnsf/ws_maxlnsf')).drop('ws_maxlnsf')
    .withColumn('ws_flnsfl',func.expr('ws_lnsfl/ws_maxlnsfl')).drop('ws_maxlnsfl')
    .withColumn('ws_c',func.expr('ws_flnc+ws_flh+ws_flhm+ws_flns+ws_flnsf+ws_flnsfl'))
  )
  df_agg_count_ln_wmaxc_t.repartition(200).write.mode("overwrite").format("parquet").save(savepath_agg_count_ln_wmaxc)



# COMMAND ----------

# MAGIC %md # author-institution

# COMMAND ----------

# in order to label the institution of each author:

# for each author in Scopus, get their latest record and select the affiliation[s] from that record.
# mod: lets take the most dominant affiliation in the last year.

tbn_df_ani_auth_affinst_nonnull="temp_df_ani_auth_affinst_nonnull"
savepath_auth_affinst_nonnull=basePath_project+"cache/"+ani['stamp']+"/"+tbn_df_ani_auth_affinst_nonnull
if file_exists(savepath_auth_affinst_nonnull):
  print("path already exists: "+savepath_auth_affinst_nonnull)
else:
  df_ani_auth_affinst_nonnull_t=(
    df_ani
    .withColumn('Au',func.explode('Au'))
    .withColumn("Au_af", func.explode("Au_af"))
    .filter('Au.Authorseq=Au_af.Authorseq') # only keep the au-af rows matching this author
    .selectExpr('*','Af[Au_af.affiliation_seq-1] as affiliation') # only keep the af rows matching this au-af.
    .drop('Af')
    .filter('affiliation IS NOT NULL') # so that we can filter to only the last publication. if it includes nulls we may select "null" as institution, not desirable.
    .filter('affiliation.affiliation_organization IS NOT NULL') # remove null org names
    .filter(func.size('affiliation.affiliation_organization')>0) # remove empty organizations
    .withColumn('auid',func.col('Au.auid').cast('long'))
    
    # select affiliations per author based on the most dominant affiliation in the last year.
    # keep only records from the final year
    .withColumn('yearRank',func.rank().over(Window.partitionBy('auid').orderBy(func.desc(sort_pub_year))))
    .filter('yearRank=1')
    # now count per author the afid occurences.
    .groupBy('auid',func.col('affiliation.afid').alias('affil_id'))
    .agg(
      func.count('*').alias('aff_occurences'),
      func.first(func.array_join('affiliation.affiliation_organization',", ")).alias('affil_name'),
      func.first('affiliation.affiliation_tag_country').alias('cntry'),
      func.max('datesort').alias('max_datesort'),
    )
    .withColumn('affRank',func.rank().over(Window.partitionBy('auid').orderBy(
      func.desc('aff_occurences'),
      func.desc('max_datesort'),
      func.asc('affil_id')
    )))
    .filter('affRank=1')
    .drop('affRank').drop('aff_occurences')

    # get name from ipr record (preferred name), otherwise default to the name as printed on the paper.
    .join(df_ipr.select(func.col('afid').alias('affil_id'),'preferred_name'),['affil_id'],'LEFT_OUTER')
    .withColumn('inst_name',func.coalesce('preferred_name','affil_name'))
  )
  print ("Generating "+savepath_auth_affinst_nonnull)
  df_ani_auth_affinst_nonnull_t.repartition(200).write.mode("overwrite").format("parquet").save(savepath_auth_affinst_nonnull)



# COMMAND ----------

# MAGIC %md # top cited selection

# COMMAND ----------

top_list_expression='(ws_ord <= 100000 OR ns_ord <= 100000) OR (rank_sm_subfield_1_ws/count_sm_subfield_1 <=.02) OR (rank_sm_subfield_1_ns/count_sm_subfield_1 <=.02)'
if file_exists(csvOutFileName_s1+'.parquet'):
  print("skip writing "+csvOutFileName_s1+'.parquet')
else:
  print("writing "+csvOutFileName_s1+'.parquet')
  
  df_agg_count_ln_wmaxc=spark.read.format("parquet").load(savepath_agg_count_ln_wmaxc)
  df_ani_auth_affinst=spark.read.format("parquet").load(savepath_auth_affinst_nonnull)
  num_should_be_empty=(
    df_ani_auth_affinst
    .withColumn('reccount',func.count('*').over(Window.partitionBy('auid')))
    .filter('reccount>1')
    .count()
  )
  print("Number of records with duplicate auid, should be 0: "+str(num_should_be_empty))
  # should be empty response.

  # complete to dataset: tie all together and calculate the rank of each author so that we can pull the top 100K.

  # ord    order sorted on c
  # author_id Scopus AUID
  # authfull  author name
  ns_wc = (Window.partitionBy(func.lit(True)).orderBy(func.desc('ns_c')))
  ws_wc = (Window.partitionBy(func.lit(True)).orderBy(func.desc('ws_c')))
  (
    df_agg_count_ln_wmaxc
    .filter('auid IS NOT NULL')
    .filter("npY1Y3 >= 2") # only rank authors >= 2 papers since 1960
    .join(spark.read.format("parquet").load(savepath_auth_name),["auid"],"LEFT_OUTER")
    .join(df_ani_auth_affinst,["auid"],"LEFT_OUTER")
    .withColumnRenamed('auid','author_id')
    .withColumn('ns_ord',func.rank().over(ns_wc))
    .withColumn('ws_ord',func.rank().over(ws_wc))
    .withColumn('self_p',(func.col('ws_ncY2Y3')-func.col('ns_ncY2Y3'))/func.col('ws_ncY2Y3'))
    .withColumn('ns_cprat',func.expr("ns_ncY2Y3/ns_ncY2Y3_cp"))
    .withColumn('ws_cprat',func.expr("ws_ncY2Y3/ws_ncY2Y3_cp"))

    .withColumn('sm_subfield_1',func.expr('IF(subfields[0] IS NULL,NULL,IF(subfields[0].SMFCMdocCountAuthor<1,NULL,subfields[0].subfield))'))
    .withColumn('sm_subfield_1_frac',func.expr('IF(subfields[0] IS NULL,NULL,IF(subfields[0].SMFCMdocCountAuthor<1,NULL,subfields[0].subfieldFrac))'))
    .withColumn('sm_subfield_2',func.expr('IF(subfields[1] IS NULL,NULL,IF(subfields[1].SMFCMdocCountAuthor<1,NULL,subfields[1].subfield))'))
    .withColumn('sm_subfield_2_frac',func.expr('IF(subfields[1] IS NULL,NULL,IF(subfields[1].SMFCMdocCountAuthor<1,NULL,subfields[1].subfieldFrac))'))
    .withColumn('sm_field',func.expr('IF(fields[0] IS NULL,NULL,IF(fields[0].SMFCMdocCountAuthor<1,NULL,fields[0].field))'))
    .withColumn('sm_field_frac',func.expr('IF(fields[0] IS NULL,NULL,IF(fields[0].SMFCMdocCountAuthor<1,NULL,fields[0].fieldFrac))'))

    # rank but don't rank if the field is none.
    .withColumn('rank_sm_subfield_1_ws',func.when(func.col('sm_subfield_1').isNull(),func.lit(None)).otherwise(func.rank().over(Window.partitionBy('sm_subfield_1').orderBy(func.desc('ws_c')))))
    .withColumn('rank_sm_subfield_1_ns',func.when(func.col('sm_subfield_1').isNull(),func.lit(None)).otherwise(func.rank().over(Window.partitionBy('sm_subfield_1').orderBy(func.desc('ns_c')))))
    .withColumn('count_sm_subfield_1',func.when(func.col('sm_subfield_1').isNull(),func.lit(None)).otherwise(func.count('*').over(Window.partitionBy('sm_subfield_1'))))

  ).repartition(200).write.mode("overwrite").format('parquet').save(csvOutFileName_s1+'.parquet')

# COMMAND ----------

# MAGIC %md ## Load cached file

# COMMAND ----------

df_agg_result_parquet=(
  spark
  .read
  .format("parquet")
  .load(csvOutFileName_s1+'.parquet')
)

# COMMAND ----------

print(ani['stamp']+" has "+str(df_agg_result_parquet.count())+" authors")
print(ani['stamp']+" has "+str(df_agg_result_parquet.filter('sm_field IS NOT NULL').count())+" authors with field assigned")# 

# COMMAND ----------

# MAGIC %md # field aggregation functions

# COMMAND ----------

from pyspark.sql import functions as func 

def authSubjectWindowColumns(window_partition,window_rank_fields,df,tableTag,column_prefix=''):
  cacheLoc=basePath_project+"cache/"+ani['stamp']+"/"+"TS3_window_"+tableTag+"_"+column_prefix+"_"+minyear+'_'+mincityear+'_'+maxyear+'_'+Y4+'_pp'+str(includepp)
  if file_exists(cacheLoc):
    print("skip writing "+cacheLoc)
  else:
    print("writing "+cacheLoc) 
    df=(
      df
      .withColumn(column_prefix+'rec_count',func.count('*').over(Window.partitionBy(window_partition)))
    )
    for field in window_rank_fields:
      # note: when ranking in descending order, the null end at the bottom. that is why for the calculation of an inverse percentile, we do not need to partition by isnull. they rank lower anyway.
      df=(
        df
        .withColumn(column_prefix+field+'_count',func.count(field).over(Window.partitionBy(window_partition)))
        .withColumn(column_prefix+field+'_rank_inv',func.rank().over(Window.partitionBy(window_partition).orderBy(func.desc(field))))
      )
    (
      df.select(*(['author_id',column_prefix+'rec_count']+[column_prefix+field+'_count' for field in window_rank_fields]+[column_prefix+field+'_rank_inv' for field in window_rank_fields]))
    ).write.save(cacheLoc)
  return spark.read.load(cacheLoc)
    

def authSubjectWindowAgg(df_result_parquet_grouped):
  return (
    df_result_parquet_grouped
    .agg(
      func.first('rec_count').alias('Auth'),
      func.count(func.expr('IF(ns_ord<=100000,TRUE,NULL)')).alias('Auth-top-100k-ns'),
      func.count(func.expr('IF(ws_ord<=100000,TRUE,NULL)')).alias('Auth-top-100k'),
      func.first('toplist_rec_count',True).alias('Auth-in-top-list'),
      # percentile cutoffs based on the upper bound of the percentile range.
      # we use the inverse-rank for that, so that the highest value is ranked 1; 
      # we subtract 1 and subtract that from the record count, divided by record count = upper bound of percentile; 
      # i.e. 1/100 = (100-(1-1))/100=1, i.e. 100%
      func.min(func.expr('IF(((ws_ncY2Y3_count-(ws_ncY2Y3_rank_inv-1))/ws_ncY2Y3_count)>=.25,ws_ncY2Y3,NULL)')).alias('Cites-25'),
      func.min(func.expr('IF(((ws_ncY2Y3_count-(ws_ncY2Y3_rank_inv-1))/ws_ncY2Y3_count)>=.50,ws_ncY2Y3,NULL)')).alias('Cites-50'),
      func.min(func.expr('IF(((ws_ncY2Y3_count-(ws_ncY2Y3_rank_inv-1))/ws_ncY2Y3_count)>=.75,ws_ncY2Y3,NULL)')).alias('Cites-75'),
      func.min(func.expr('IF(((ws_ncY2Y3_count-(ws_ncY2Y3_rank_inv-1))/ws_ncY2Y3_count)>=.90,ws_ncY2Y3,NULL)')).alias('Cites-90'),
      func.min(func.expr('IF(((ws_ncY2Y3_count-(ws_ncY2Y3_rank_inv-1))/ws_ncY2Y3_count)>=.95,ws_ncY2Y3,NULL)')).alias('Cites-95'),
      func.min(func.expr('IF(((ws_ncY2Y3_count-(ws_ncY2Y3_rank_inv-1))/ws_ncY2Y3_count)>=.99,ws_ncY2Y3,NULL)')).alias('Cites-99'),
      func.min(func.expr('IF(((ws_c_count-(ws_c_rank_inv-1))/ws_c_count)>=.25,ws_c,NULL)')).alias('c-25'),
      func.min(func.expr('IF(((ws_c_count-(ws_c_rank_inv-1))/ws_c_count)>=.50,ws_c,NULL)')).alias('c-50'),
      func.min(func.expr('IF(((ws_c_count-(ws_c_rank_inv-1))/ws_c_count)>=.75,ws_c,NULL)')).alias('c-75'),
      func.min(func.expr('IF(((ws_c_count-(ws_c_rank_inv-1))/ws_c_count)>=.90,ws_c,NULL)')).alias('c-90'),
      func.min(func.expr('IF(((ws_c_count-(ws_c_rank_inv-1))/ws_c_count)>=.95,ws_c,NULL)')).alias('c-95'),
      func.min(func.expr('IF(((ws_c_count-(ws_c_rank_inv-1))/ws_c_count)>=.99,ws_c,NULL)')).alias('c-99'),

      func.min(func.expr('IF(((self_p_count-(self_p_rank_inv-1))/self_p_count)>=.95,`self%`,NULL)')).alias('selfp-95'),
      func.min(func.expr('IF(((self_p_count-(self_p_rank_inv-1))/self_p_count)>=.99,`self%`,NULL)')).alias('selfp-99'),
      func.min(func.expr('IF(((ws_cprat_count-(ws_cprat_rank_inv-1))/ws_cprat_count)>=.95,ws_cprat,NULL)')).alias('cprat-95'),
      func.min(func.expr('IF(((ws_cprat_count-(ws_cprat_rank_inv-1))/ws_cprat_count)>=.99,ws_cprat,NULL)')).alias('cprat-99'),
      func.min(func.expr('IF(((ns_cprat_count-(ns_cprat_rank_inv-1))/ns_cprat_count)>=.95,ws_cprat,NULL)')).alias('cprat-ns-95'),
      func.min(func.expr('IF(((ns_cprat_count-(ns_cprat_rank_inv-1))/ns_cprat_count)>=.99,ns_cprat,NULL)')).alias('cprat-ns-99'),    

      func.min(func.expr('IF(((toplist_self_p_count-(toplist_self_p_rank_inv-1))/toplist_self_p_count)>=.95,`self%`,NULL)')).alias('top-list-selfp-95'),
      func.min(func.expr('IF(((toplist_self_p_count-(toplist_self_p_rank_inv-1))/toplist_self_p_count)>=.99,`self%`,NULL)')).alias('top-list-selfp-99'),
      func.min(func.expr('IF(((toplist_ws_cprat_count-(toplist_ws_cprat_rank_inv-1))/toplist_ws_cprat_count)>=.95,ws_cprat,NULL)')).alias('top-list-cprat-95'),
      func.min(func.expr('IF(((toplist_ws_cprat_count-(toplist_ws_cprat_rank_inv-1))/toplist_ws_cprat_count)>=.99,ws_cprat,NULL)')).alias('top-list-cprat-99'),
      func.min(func.expr('IF(((toplist_ns_cprat_count-(toplist_ns_cprat_rank_inv-1))/toplist_ns_cprat_count)>=.95,ws_cprat,NULL)')).alias('top-list-cprat-ns-95'),
      func.min(func.expr('IF(((toplist_ns_cprat_count-(toplist_ns_cprat_rank_inv-1))/toplist_ns_cprat_count)>=.99,ns_cprat,NULL)')).alias('top-list-cprat-ns-99'),      
      
      # count records in top %, which be around <percentile>-of-top-2% due to the selection criteria for top-list.
      func.count(func.expr('IF(toplist_self_p_count IS NOT NULL,IF(((toplist_self_p_count-(toplist_self_p_rank_inv-1))/toplist_self_p_count)>=.95,author_id,NULL),NULL)')).alias('Auth-in-top-list-selfp-95'),
      func.count(func.expr('IF(toplist_self_p_count IS NOT NULL,IF(((toplist_self_p_count-(toplist_self_p_rank_inv-1))/toplist_self_p_count)>=.99,author_id,NULL),NULL)')).alias('Auth-in-top-list-selfp-99'),
      func.count(func.expr('IF(toplist_ws_cprat_count IS NOT NULL,IF(((toplist_ws_cprat_count-(toplist_ws_cprat_rank_inv-1))/toplist_ws_cprat_count)>=.95,author_id,NULL),NULL)')).alias('Auth-in-top-list-cprat-95'),
      func.count(func.expr('IF(toplist_ws_cprat_count IS NOT NULL,IF(((toplist_ws_cprat_count-(toplist_ws_cprat_rank_inv-1))/toplist_ws_cprat_count)>=.99,author_id,NULL),NULL)')).alias('Auth-in-top-list-cprat-99'),
      func.count(func.expr('IF(toplist_ns_cprat_count IS NOT NULL,IF(((toplist_ns_cprat_count-(toplist_ns_cprat_rank_inv-1))/toplist_ns_cprat_count)>=.95,author_id,NULL),NULL)')).alias('Auth-in-top-list-cprat-ns-95'),
      func.count(func.expr('IF(toplist_ns_cprat_count IS NOT NULL,IF(((toplist_ns_cprat_count-(toplist_ns_cprat_rank_inv-1))/toplist_ns_cprat_count)>=.99,author_id,NULL),NULL)')).alias('Auth-in-top-list-cprat-ns-99'),
      
    )
  )

def windowedTableS3(df,window_partition,tableTag):
  cacheLoc=basePath_project+"cache/"+ani['stamp']+"/"+"TS3_grouped_"+tableTag+"_"+minyear+'_'+mincityear+'_'+maxyear+'_'+Y4+'_pp'+str(includepp)
  if file_exists(cacheLoc):
    print("skip writing "+cacheLoc)
  else:
    print("writing "+cacheLoc) 
    # columns we need ranked window columns for, for percentile extraction:
    window_rank_fields=[
      'ws_ncY2Y3','ws_c','self_p','ws_cprat','ns_cprat',
    ]
    # columns we need ranked window columns for, for percentile extraction of the top-list subset:
    window_rank_fields_toplist=[
      'self_p','ws_cprat','ns_cprat',
    ]
    authSubjectWindowAgg(
      df
      .join(
        (authSubjectWindowColumns(window_partition,window_rank_fields,df,tableTag)),
        ['author_id']
      )
      .join(
        (authSubjectWindowColumns(window_partition,window_rank_fields_toplist,df.filter(top_list_expression),tableTag,'toplist_')),
        ['author_id'],'LEFT_OUTER'
      )
      .withColumn('self%',func.expr("CONCAT(ROUND(100*self_p,2),' %')"))
      .groupBy(window_partition.alias('subject_column'))
    ).write.save(cacheLoc)
  return spark.read.load(cacheLoc)
    




# COMMAND ----------

# MAGIC %md ## field table

# COMMAND ----------

# table S3 by field (20)
if file_exists(csvOutFileName_s3+'_SM_FIELD.csv'):
  print("skip writing "+csvOutFileName_s3+'_SM_FIELD.csv')
else:
  print("writing "+csvOutFileName_s3+'_SM_FIELD.csv')
  (
    windowedTableS3(df_agg_result_parquet,func.col('sm_field'),'SM_FIELD')
    .union(
      windowedTableS3(df_agg_result_parquet,func.lit('TOTAL'),'TOTAL')
    )
  ).repartition(1).write.mode("overwrite").csv(csvOutFileName_s3+'_SM_FIELD.csv',header = 'true', quote='"', escape='"')




# COMMAND ----------

# MAGIC %md ## subfield table

# COMMAND ----------

# table s3 by subfield (174)
if file_exists(csvOutFileName_s3+'_SM_SUBFIELD.csv'):
  print("skip writing "+csvOutFileName_s3+'_SM_SUBFIELD.csv')
else:
  print("writing "+csvOutFileName_s3+'_SM_SUBFIELD.csv')
  (
    windowedTableS3(df_agg_result_parquet,func.col('sm_subfield_1'),'SM_SUBFIELD')
    .union(
      windowedTableS3(df_agg_result_parquet,func.lit('TOTAL'),'TOTAL')
    )
  ).repartition(1).write.mode("overwrite").csv(csvOutFileName_s3+'_SM_SUBFIELD.csv',header = 'true', quote='"', escape='"')


# COMMAND ----------

# MAGIC %md # main table selective

# COMMAND ----------

from pyspark.sql.window import Window

def select_fields_top_list_output(df):
  return (
    df
    .select(
      "author_id",
      func.regexp_replace("authfull", '"', '').alias("authfull"),
      func.regexp_replace("inst_name", '"', '').alias("inst_name"),
      "cntry",
      func.col("npY1Y3").alias('np'+Y1+Y3),
      func.col('firstPubYear').alias('firstyr'),
      func.col('lastPubYear').alias('lastyr'),
      func.col("ns_ord").alias('rank (ns)'),
      func.col("ns_ncY2Y3").alias('nc'+Y2+Y3+' (ns)'),
      func.col("ns_hY3").alias('h'+Y3+' (ns)'),
      func.col("ns_hmY3").alias('hm'+Y3+' (ns)'),
      func.col("ns_nps").alias('nps (ns)'),
      func.col("ns_ncs").alias('ncs (ns)'),
      func.col("ns_npsf").alias('cpsf (ns)'),
      func.col("ns_ncsf").alias('ncsf (ns)'),
      func.col("ns_npsfl").alias('npsfl (ns)'),
      func.col("ns_ncsfl").alias('ncsfl (ns)'),

      func.col("ns_c").alias('c (ns)'),
      func.col("ns_ncY2Y3_cp").alias('npciting (ns)'),
      func.col("ns_cprat").alias('cprat (ns)'),  
      func.col('ns_npcY1Y3').alias('np'+Y1+Y3+' cited'+Y2+Y3+' (ns)'),
      func.expr("CONCAT(ROUND(100*self_p,2),' %')").alias('self%'),

      func.col("ws_ord").alias('rank'),
      func.col("ws_ncY2Y3").alias('nc'+Y2+Y3),
      func.col("ws_hY3").alias('h'+Y3),
      func.col("ws_hmY3").alias('hm'+Y3),
      func.col("ws_nps").alias('nps'),
      func.col("ws_ncs").alias('ncs'),
      func.col("ws_npsf").alias('cpsf'),
      func.col("ws_ncsf").alias('ncsf'),
      func.col("ws_npsfl").alias('npsfl'),
      func.col("ws_ncsfl").alias('ncsfl'),    

      func.col("ws_c").alias('c'),
      func.col("ws_ncY2Y3_cp").alias('npciting'),
      func.col('ws_cprat').alias('cprat'),
      func.col('ws_npcY1Y3').alias('np'+Y1+Y3+' cited'+Y2+Y3),

      func.col("npY1Y3_d").alias('np'+Y1+Y3+'_d'),
      func.col("ws_ncY2Y3_d").alias('nc'+Y2+Y3+'_d'),

      func.col('sm_subfield_1').alias('sm-subfield-1'),
      func.col('sm_subfield_1_frac').alias('sm-subfield-1-frac'),
      func.col('sm_subfield_2').alias('sm-subfield-2'),
      func.col('sm_subfield_2_frac').alias('sm-subfield-2-frac'),
      func.col('sm_field').alias('sm-field'),
      func.col('sm_field_frac').alias('sm-field-frac'),   
      func.col('rank_sm_subfield_1_ws').alias('rank sm-subfield-1'),
      func.col('rank_sm_subfield_1_ns').alias('rank sm-subfield-1 (ns)'),
      func.col('count_sm_subfield_1').alias('sm-subfield-1 count'),

    )
    .orderBy(func.asc('rank (ns)'))
  )

df_agg_result_csv_top2p_bysubfield_100K_combined=(
  select_fields_top_list_output(
    df_agg_result_parquet
    .filter(top_list_expression)
  )
)
if file_exists(csvOutFileName_s1+'_top2p_bysubfield_100K_combined.csv'):
  print("skip writing "+csvOutFileName_s1+'_top2p_bysubfield_100K_combined.csv')
else:
  print("writing "+csvOutFileName_s1+'_top2p_bysubfield_100K_combined.csv')
  df_agg_result_csv_top2p_bysubfield_100K_combined.repartition(1).write.mode("overwrite").csv(csvOutFileName_s1+'_top2p_bysubfield_100K_combined.csv',header = 'true')


# COMMAND ----------

# MAGIC %md # Excel exports

# COMMAND ----------

import openpyxl
import os
from shutil import copyfile

# COMMAND ----------

# MAGIC %md ## Table 1

# COMMAND ----------

f_orange=openpyxl.styles.PatternFill(patternType='solid',fill_type='solid',fgColor=openpyxl.styles.Color('FFC000'))
f_green=openpyxl.styles.PatternFill(patternType='solid',fill_type='solid',fgColor=openpyxl.styles.Color('92D050'))
f_yellow=openpyxl.styles.PatternFill(patternType='solid',fill_type='solid',fgColor=openpyxl.styles.Color('FFFF00'))
data_cols_styles=[
  None,
  None,
  None,
  ['int',None],
  ['rawint',None],
  ['rawint',None],
  ['int',f_orange],
  ['int',f_orange],
  ['int',f_orange],
  ['dec',f_orange],
  ['int',f_orange],
  ['int',f_orange],
  ['int',f_orange],
  ['int',f_orange],
  ['int',f_orange],
  ['int',f_orange],
  ['dec',f_orange],
  ['int',f_green],
  ['dec',f_green],
  ['int',f_green],
  ['perc',f_green],
  ['int',f_yellow],
  ['int',f_yellow],
  ['int',f_yellow],
  ['dec',f_yellow],
  ['int',f_yellow],
  ['int',f_yellow],
  ['int',f_yellow],
  ['int',f_yellow],
  ['int',f_yellow],
  ['int',f_yellow],
  ['dec',f_yellow],  
  ['int',f_green],
  ['dec',f_green],
  ['int',f_green],  
  ['int',None],
  ['int',None],
  None,
  ['dec',None],
  None,
  ['dec',None],
  None,
  ['dec',None],
  ['int',None],
  ['int',None],
  ['int',None],
]

# return value as float, optionally specify what to do if null
def floatVal(x,nullRet=''):
  return nullRet if x is None else float(x)
def intVal(x,nullRet=''):
  return nullRet if x is None else int(x)
def perc2fracVal(x,nullRet='N/A'):
  return nullRet if x is None else float(x[0:-1])/float(100)

def wsl_fill_sheet_data(df_export,wsl,printheaders=True):
  df_collected=df_export.collect()
  column_headers = df_export.schema.names
  if (printheaders):
    # header line
    for col0,column_header in enumerate(column_headers):
      cell=wsl.cell(column=col0+1, row=1, value=column_header)
      cell.font=openpyxl.styles.Font(color='000000', bold=True)
      if data_cols_styles[col0]!=None:
        if data_cols_styles[col0][1]!=None:
          cell.fill=data_cols_styles[col0][1]
      

  # data
  for row0,row in enumerate(df_collected):
    for col0,column_header in enumerate(column_headers):
      # need to +2 for row0 as we also have a header row.
      cell=wsl.cell(column=col0+1, row=row0+2, value=row[column_header])
      if data_cols_styles[col0]!=None:
        if data_cols_styles[col0][0]=='dec':
          cell.number_format = '#,##0.0000' 
          cell.value=floatVal(row[column_header])
        elif data_cols_styles[col0][0]=='int':
          cell.number_format = '#,##0' 
          cell.value=intVal(row[column_header])
        elif data_cols_styles[col0][0]=='rawint':
          cell.value=intVal(row[column_header])
        elif data_cols_styles[col0][0]=='perc':
          cell.number_format = '0.00%' 
          cell.value=perc2fracVal(row[column_header])
        if data_cols_styles[col0][1]!=None:
          cell.fill=data_cols_styles[col0][1]                                             
  wsl.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 22
  wsl.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 22
  wsl.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 22          
  return wsl


def wsl_fill_sheet_key(df_export,wsl,printheaders=True):
  df_collected=df_export.collect()
  column_headers = df_export.schema.names
  if (printheaders):
    # header line
    for col0,column_header in enumerate(column_headers):
      cell=wsl.cell(column=col0+1, row=1, value=column_header)
      cell.font=openpyxl.styles.Font(color='FF0000', bold=True)

  # data
  for row0,row in enumerate(df_collected):
    for col0,column_header in enumerate(column_headers):
      # need to +2 for row0 as we also have a header row.
      cell=wsl.cell(column=col0+1, row=row0+2, value=row[column_header])
      if col0==0:
        cell.font=openpyxl.styles.Font(color='000000', bold=True)
        if data_cols_styles[row0]!=None:
          if data_cols_styles[row0][1]!=None:
            cell.fill=data_cols_styles[row0][1]          
  wsl.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 22
  wsl.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 22
  wsl.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 22
  return wsl


key_data=[
#['FIELD','BASIS','DESCRIPTION'],
['authfull','','author name'],
['inst_name','','institution name (large institutions only)'],
['cntry','','country associated with most recent institution'],
[f'np{Y1}{Y3}','',f'# papers {minyear}-{maxyear}'],
['firstyr','','year of first publication'],
['lastyr','','year of most recent publication'],
['rank (ns)','self-citations excluded','rank based on composite score c'],
[f'nc{Y2}{Y3} (ns)','self-citations excluded',f'total cites {mincityear}-{maxyear}'],
[f'h{Y3} (ns)','self-citations excluded',f'h-index as of end-{maxyear}'],
[f'hm{Y3} (ns)','self-citations excluded',f'hm-index as of end-{maxyear}'],
['nps (ns)','self-citations excluded','number of single authored papers'],
['ncs (ns)','self-citations excluded','total cites to single authored papers'],
['cpsf (ns)','self-citations excluded','number of single+first authored papers'],
['ncsf (ns)','self-citations excluded','total cites to single+first authored papers'],
['npsfl (ns)','self-citations excluded','number of single+first+last authored papers'],
['ncsfl (ns)','self-citations excluded','total cites to single+first+last authored papers'],
['c (ns)','self-citations excluded','composite score'],
['npciting (ns)','self-citations excluded','number of distinct citing papers'],
['cprat (ns)','self-citations excluded','ratio of total citations to distinct citing papers'],
[f'np{Y1}{Y3} cited{Y2}{Y3} (ns)','self-citations excluded',f'number of papers {minyear}-{maxyear} that have been cited at least once'],
['self%','','self-citation percentage'],
['rank','all citations','rank based on composite score c'],
[f'nc{Y2}{Y3}','all citations',f'total cites {mincityear}-{maxyear}'],
[f'h{Y3}','all citations',f'h-index as of end-{maxyear}'],
[f'hm{Y3}','all citations',f'hm-index as of end-{maxyear}'],
['nps','all citations','number of single authored papers'],
['ncs','all citations','total cites to single authored papers'],
['cpsf','all citations','number of single+first authored papers'],
['ncsf','all citations','total cites to single+first authored papers'],
['npsfl','all citations','number of single+first+last authored papers'],
['ncsfl','all citations','total cites to single+first+last authored papers'],
['c','all citations','composite score'],
['npciting','all citations','number of distinct citing papers'],
['cprat','all citations','ratio of total citations to distinct citing papers'],
[f'np{Y1}{Y3} cited{Y2}{Y3}','all citations',f'number of papers {minyear}-{maxyear} that have been cited at least once'],
[f'np{Y1}{Y3}_d','',f'# papers {minyear}-{maxyear} in titles that are discontinued in Scopus'],
[f'nc{Y2}{Y3}_d','',f'total cites {mincityear}-{maxyear} from titles that are discontinued in Scopus'],
['sm-subfield-1','all citations','top ranked Science-Metrix category (subfield) for author'],
['sm-subfield-1-frac','all citations','associated category fraction'],
['sm-subfield-2','all citations','second ranked Science-Metrix category (subfield) for author'],
['sm-subfield-2-frac','all citations','associated category fraction'],
['sm-field','all citations','top ranked higher-level Science-Metrix category (field) for author'],
['sm-field-frac','all citations','associated category fraction'],
['rank sm-subfield-1','all citations','rank of c within category sm-subfield-1'],
['rank sm-subfield-1 (ns)','self-citations excluded','rank of c (ns) within category sm-subfield-1'],
['sm-subfield-1 count','','total number of authors within category sm-subfield-1'],
]

def save_wb(wb,fn_base):
  # save to local store (can't do a random write to dbfs)
  wb.save(f'/tmp/{fn_base}.xlsx')
  # copy to dbfs for http download link
  copyfile(f'/tmp/{fn_base}.xlsx', f'/dbfs/FileStore/{fn_base}.xlsx')
  # also copy to our s3 bucket
  s3target=os.path.join(basePath_project,'excel_export',f'{fn_base}.xlsx')
  dbutils.fs.cp(f'dbfs:/FileStore/{fn_base}.xlsx',s3target)
  print(f'stored for project: {s3target}')
           
  return f'https://elsevier.cloud.databricks.com/files/{fn_base}.xlsx'  

def gen_workbook_table1(fn_base,df_a,key_data):
  wb = openpyxl.Workbook()
  
  # Sheet 1
  wsl =  wb.active # first sheet to add is always already there.
  wsl.title = "Key"
  
  wsl = wsl_fill_sheet_key((spark.createDataFrame(list(map(lambda x: Row(FIELD=x[0],BASIS=x[1],DESCRIPTION=x[2]), key_data)))),wsl,True)
  
  wsl = wsl_fill_sheet_data(df_a.drop('author_id'),wb.create_sheet("Data"))

  # save the wb, return url.
  return save_wb(wb,fn_base)



# COMMAND ----------

# MAGIC %md ## Table 2

# COMMAND ----------

data_cols_width_styles_table2=[
  [10,'int'],
  [17.57,'int'],
  [13,'perc'],
  [14,'int'],
  [8.43,'perc'],
  [14.71,'int'],
  [10,'perc'],
  [8.57,'int'],
  [8.57,'int'],
  [8.57,'int'],
  [8.57,'int'],
  [8.57,'int'],
  [8.57,'int'],
  [5.90,'dec'],
  [5.90,'dec'],
  [5.90,'dec'],
  [5.90,'dec'],
  [5.90,'dec'],
  [5.90,'dec'],
  [8.43,'perc'],
  [8.43,'perc'],
  [8.43,'dec'],
  [8.43,'dec'],
  [8.43,'dec'],
  [8.43,'dec'],
  [8.43,'perc'],
]
data_cols_width_styles_table2_field=[
  [25,None],
  [40,None],
]+data_cols_width_styles_table2
data_cols_width_styles_table2_subfield=[
  [25,None],
  [40,None],
  [41.86,None],# subfield  
]+data_cols_width_styles_table2


def wsl_fill_sheet_data_fieldsubfield(df_export,wsl,data_cols_width_styles,printheaders=True):
  df_collected=df_export.collect()
  column_headers = df_export.schema.names
  if (printheaders):
    # header line
    for col0,column_header in enumerate(column_headers):
      cell=wsl.cell(column=col0+1, row=1, value=column_header)
      cell.font=openpyxl.styles.Font(color='000000', bold=True)

  # data
  for row0,row in enumerate(df_collected):
    for col0,column_header in enumerate(column_headers):
      # need to +2 for row0 as we also have a header row.
      cell=wsl.cell(column=col0+1, row=row0+2, value=row[column_header])
      if (row[column_header]!=None):
        if data_cols_width_styles[col0][1]!=None:
          if data_cols_width_styles[col0][1]=='dec':
            cell.number_format = '#,##0.000' 
            cell.value=float(row[column_header])
          elif data_cols_width_styles[col0][1]=='int':
            cell.number_format = '#,##0' 
            cell.value=int(row[column_header])
          elif data_cols_width_styles[col0][1]=='rawint':
            cell.value=int(row[column_header])
          elif data_cols_width_styles[col0][1]=='perc':
            cell.number_format = '0.00%' 
            cell.value=float(row[column_header][0:-1])/float(100)
          elif data_cols_width_styles[col0][1]=='perc_r':
            cell.number_format = '0.00%' 
            cell.value=float(row[column_header])
  for col0,width in enumerate(data_cols_width_styles):
    wsl.column_dimensions[openpyxl.utils.get_column_letter(col0+1)].width = width[0]
  return wsl


def gen_workbook_table2_field_subfield(fn_base,df_table_2_field,df_table_2_subfield,data_cols_width_styles_table2_field,data_cols_width_styles_table2_subfield):
  wb = openpyxl.Workbook()
  
#   fn_base=os.path.basename(csvOutFileName_s1[9:]+'_field_subfield_thresholds')
  
  # Sheet 1
  wsl =  wb.active # first sheet to add is always already there.
  wsl.title = "Field"
  print('Field')
  wsl = wsl_fill_sheet_data_fieldsubfield(df_table_2_field,wsl,data_cols_width_styles_table2_field)
  print('SubField')
  wsl = wsl_fill_sheet_data_fieldsubfield(df_table_2_subfield,wb.create_sheet("SubField"),data_cols_width_styles_table2_subfield)

  # save the wb, return url.
  return save_wb(wb,fn_base)



# COMMAND ----------

# MAGIC %md ## Table 3

# COMMAND ----------

data_cols_width_styles_table3=[
  [10,'dec'],
  [10,'dec'],
  [10,'dec'],
  [10,'dec'],
  [10,'dec'],
  [10,'dec'],
  [10,'dec'],
  [10,'dec'],
  [10,'dec'],
  [10,'dec'],
  [10,'dec'],
  [10,'dec'],
]

def wsl_fill_sheet_data_maxvals(df_export,wsl,data_cols_width_styles,printheaders=True):
  df_collected=df_export.collect()
  column_headers = df_export.schema.names
  if (printheaders):
    # header line
    for col0,column_header in enumerate(column_headers):
      cell=wsl.cell(column=col0+1, row=1, value=column_header)
      cell.font=openpyxl.styles.Font(color='000000', bold=True)

  # data
  for row0,row in enumerate(df_collected):
    for col0,column_header in enumerate(column_headers):
      # need to +2 for row0 as we also have a header row.
      cell=wsl.cell(column=col0+1, row=row0+2, value=row[column_header])
      if (row[column_header]!=None):
        if data_cols_width_styles[col0][1]!=None:
          if data_cols_width_styles[col0][1]=='dec':
            cell.number_format = '#,##0.000' 
            cell.value=float(row[column_header])
          elif data_cols_width_styles[col0][1]=='int':
            cell.number_format = '#,##0' 
            cell.value=int(row[column_header])
          elif data_cols_width_styles[col0][1]=='rawint':
            cell.value=int(row[column_header])
          elif data_cols_width_styles[col0][1]=='perc':
            cell.number_format = '0.00%' 
            cell.value=float(row[column_header][0:-1])/float(100)
          elif data_cols_width_styles[col0][1]=='perc_r':
            cell.number_format = '0.00%' 
            cell.value=float(row[column_header])
  for col0,width in enumerate(data_cols_width_styles):
    wsl.column_dimensions[openpyxl.utils.get_column_letter(col0+1)].width = width[0]
  return wsl

def gen_workbook_table3_maxvals(fn_base,data_cols_width_styles_table3):
  wb = openpyxl.Workbook()
  
  # Sheet 1
  wsl =  wb.active # first sheet to add is always already there.
  wsl.title = "Maxvalues"
  wsl = wsl_fill_sheet_data_maxvals(df_table_3_maxvalues,wsl,data_cols_width_styles_table3)

  # save the wb, return url.
  return save_wb(wb,fn_base)


